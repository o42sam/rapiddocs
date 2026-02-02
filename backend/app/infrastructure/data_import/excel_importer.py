"""
Excel Data Importer Implementation.
Implements IDataImporter interface for Excel files.
"""

from typing import List, Dict, Any, Optional
from pathlib import Path

from ...domain.interfaces.data_importer import IDataImporter
from ...shared.logger import get_logger

logger = get_logger("excel_importer")


class ExcelImporter(IDataImporter):
    """
    Excel implementation of data importer.
    Imports statistics and other data from Excel files (.xlsx, .xls).

    Requires: openpyxl for .xlsx, xlrd for .xls

    Expected columns for infographic:
    - name: Statistic name
    - value: Numeric value
    - unit: Unit of measurement
    - visualization_type: Chart type
    - category: Optional category
    - description: Optional description
    """

    # Column aliases (same as CSV importer)
    COLUMN_ALIASES = {
        'name': ['name', 'label', 'title', 'statistic', 'stat_name'],
        'value': ['value', 'amount', 'number', 'quantity', 'stat_value'],
        'unit': ['unit', 'units', 'measurement', 'uom'],
        'visualization_type': ['visualization_type', 'viz_type', 'chart_type', 'chart', 'type'],
        'category': ['category', 'group', 'cat'],
        'description': ['description', 'desc', 'details', 'notes']
    }

    def __init__(self):
        """Initialize the Excel importer."""
        self._pandas_available = False
        self._openpyxl_available = False

        # Check for pandas
        try:
            import pandas as pd
            self._pandas_available = True
            logger.info("Excel Importer initialized with pandas support")
        except ImportError:
            logger.warning("pandas not available, using fallback")

        # Check for openpyxl
        try:
            import openpyxl
            self._openpyxl_available = True
        except ImportError:
            logger.warning("openpyxl not available")

    @property
    def supported_extensions(self) -> List[str]:
        """Return list of supported file extensions."""
        return ['.xlsx', '.xls']

    async def import_file(
        self,
        file_path: Path,
        options: Optional[Dict[str, Any]] = None
    ) -> List[Dict[str, Any]]:
        """
        Import data from Excel file.

        Args:
            file_path: Path to the Excel file
            options: Optional import options
                - sheet_name: Sheet name or index (default: first sheet)
                - skip_rows: Number of rows to skip
                - header_row: Row index containing headers (default: 0)

        Returns:
            List of dictionaries containing imported data
        """
        logger.info(f"Importing Excel file: {file_path}")

        options = options or {}
        sheet_name = options.get('sheet_name', 0)
        skip_rows = options.get('skip_rows', 0)

        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"Excel file not found: {file_path}")

        if self._pandas_available:
            return await self._import_with_pandas(file_path, sheet_name, skip_rows)
        elif self._openpyxl_available:
            return await self._import_with_openpyxl(file_path, sheet_name, skip_rows)
        else:
            raise ImportError("Neither pandas nor openpyxl is available for Excel import")

    async def _import_with_pandas(
        self,
        file_path: Path,
        sheet_name: Any,
        skip_rows: int
    ) -> List[Dict[str, Any]]:
        """Import using pandas."""
        import pandas as pd

        try:
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name,
                skiprows=skip_rows
            )

            # Create column mapping
            column_mapping = self._create_column_mapping(df.columns.tolist())
            logger.debug(f"Column mapping: {column_mapping}")

            data = []
            for _, row in df.iterrows():
                try:
                    mapped_row = self._map_row_pandas(row, column_mapping)
                    if self._validate_row(mapped_row):
                        data.append(mapped_row)
                except Exception as e:
                    logger.warning(f"Skipping invalid row: {e}")
                    continue

            logger.info(f"Imported {len(data)} records from Excel")
            return data

        except Exception as e:
            logger.error(f"Failed to import Excel with pandas: {e}")
            raise

    async def _import_with_openpyxl(
        self,
        file_path: Path,
        sheet_name: Any,
        skip_rows: int
    ) -> List[Dict[str, Any]]:
        """Import using openpyxl directly."""
        import openpyxl

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Get sheet
            if isinstance(sheet_name, int):
                ws = wb.worksheets[sheet_name]
            else:
                ws = wb[sheet_name]

            rows = list(ws.iter_rows(min_row=skip_rows + 1, values_only=True))

            if not rows:
                return []

            # First row is headers
            headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[0])]
            column_mapping = self._create_column_mapping(headers)

            data = []
            for row in rows[1:]:
                try:
                    row_dict = dict(zip(headers, row))
                    mapped_row = self._map_row_dict(row_dict, column_mapping)
                    if self._validate_row(mapped_row):
                        data.append(mapped_row)
                except Exception as e:
                    logger.warning(f"Skipping invalid row: {e}")
                    continue

            wb.close()

            logger.info(f"Imported {len(data)} records from Excel")
            return data

        except Exception as e:
            logger.error(f"Failed to import Excel with openpyxl: {e}")
            raise

    async def validate_file(
        self,
        file_path: Path,
        expected_columns: Optional[List[str]] = None
    ) -> bool:
        """
        Validate Excel file format and structure.

        Args:
            file_path: Path to the file to validate
            expected_columns: Optional list of expected column names

        Returns:
            True if file is valid, False otherwise
        """
        logger.info(f"Validating Excel file: {file_path}")

        if not file_path.exists():
            logger.error("File does not exist")
            return False

        if file_path.suffix.lower() not in self.supported_extensions:
            logger.error(f"Invalid file extension: {file_path.suffix}")
            return False

        try:
            if self._pandas_available:
                import pandas as pd
                df = pd.read_excel(file_path, nrows=1)
                headers = df.columns.tolist()
            elif self._openpyxl_available:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active
                headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
                wb.close()
            else:
                logger.error("No Excel library available")
                return False

            # Validate expected columns
            if expected_columns:
                column_mapping = self._create_column_mapping(headers)
                for expected in expected_columns:
                    if expected not in column_mapping:
                        logger.warning(f"Missing expected column: {expected}")
                        return False

            logger.info("Excel file validation passed")
            return True

        except Exception as e:
            logger.error(f"Excel validation failed: {e}")
            return False

    def preview(
        self,
        file_path: Path,
        rows: int = 5
    ) -> List[Dict[str, Any]]:
        """
        Preview first N rows of Excel data.

        Args:
            file_path: Path to the file to preview
            rows: Number of rows to preview

        Returns:
            List of preview data dictionaries
        """
        logger.info(f"Previewing Excel file: {file_path} (first {rows} rows)")

        if not file_path.exists():
            return []

        try:
            if self._pandas_available:
                import pandas as pd
                df = pd.read_excel(file_path, nrows=rows)
                column_mapping = self._create_column_mapping(df.columns.tolist())

                preview_data = []
                for _, row in df.iterrows():
                    try:
                        mapped_row = self._map_row_pandas(row, column_mapping)
                        preview_data.append(mapped_row)
                    except Exception:
                        continue

                return preview_data

            elif self._openpyxl_available:
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                ws = wb.active

                all_rows = list(ws.iter_rows(max_row=rows + 1, values_only=True))
                if not all_rows:
                    return []

                headers = [str(h) if h else f'col_{i}' for i, h in enumerate(all_rows[0])]
                column_mapping = self._create_column_mapping(headers)

                preview_data = []
                for row in all_rows[1:]:
                    try:
                        row_dict = dict(zip(headers, row))
                        mapped_row = self._map_row_dict(row_dict, column_mapping)
                        preview_data.append(mapped_row)
                    except Exception:
                        continue

                wb.close()
                return preview_data

            return []

        except Exception as e:
            logger.error(f"Preview failed: {e}")
            return []

    def _create_column_mapping(self, headers: List[str]) -> Dict[str, str]:
        """Create mapping from standard names to actual column names."""
        mapping = {}
        headers_lower = {str(h).lower().strip(): str(h) for h in headers if h}

        for standard_name, aliases in self.COLUMN_ALIASES.items():
            for alias in aliases:
                if alias in headers_lower:
                    mapping[standard_name] = headers_lower[alias]
                    break

        return mapping

    def _map_row_pandas(self, row, column_mapping: Dict[str, str]) -> Dict[str, Any]:
        """Map a pandas row to standard column names."""
        mapped = {}

        for standard_name, actual_name in column_mapping.items():
            value = row.get(actual_name)

            if standard_name == 'value':
                try:
                    import pandas as pd
                    if pd.isna(value):
                        mapped[standard_name] = 0.0
                    else:
                        mapped[standard_name] = float(value)
                except (ValueError, TypeError):
                    mapped[standard_name] = 0.0
            else:
                import pandas as pd
                if pd.isna(value):
                    mapped[standard_name] = ''
                else:
                    mapped[standard_name] = str(value).strip()

        # Set defaults
        self._set_defaults(mapped)
        return mapped

    def _map_row_dict(self, row: Dict[str, Any], column_mapping: Dict[str, str]) -> Dict[str, Any]:
        """Map a dictionary row to standard column names."""
        mapped = {}

        for standard_name, actual_name in column_mapping.items():
            value = row.get(actual_name)

            if standard_name == 'value':
                try:
                    if value is None:
                        mapped[standard_name] = 0.0
                    else:
                        mapped[standard_name] = float(value)
                except (ValueError, TypeError):
                    mapped[standard_name] = 0.0
            else:
                mapped[standard_name] = str(value).strip() if value else ''

        self._set_defaults(mapped)
        return mapped

    def _set_defaults(self, mapped: Dict[str, Any]) -> None:
        """Set default values for missing fields."""
        if 'name' not in mapped or not mapped['name']:
            mapped['name'] = 'Unnamed'
        if 'value' not in mapped:
            mapped['value'] = 0.0
        if 'unit' not in mapped or not mapped['unit']:
            mapped['unit'] = 'units'
        if 'visualization_type' not in mapped or not mapped['visualization_type']:
            mapped['visualization_type'] = 'bar_chart'

    def _validate_row(self, row: Dict[str, Any]) -> bool:
        """Validate a mapped row."""
        if not row.get('name'):
            return False

        try:
            float(row.get('value', 0))
        except (ValueError, TypeError):
            return False

        valid_types = ['bar_chart', 'line_chart', 'pie_chart', 'gauge_chart', 'number']
        viz_type = row.get('visualization_type', 'bar_chart').lower()
        if viz_type not in valid_types:
            row['visualization_type'] = 'bar_chart'

        return True

    def get_status(self) -> Dict[str, Any]:
        """Get importer status information."""
        return {
            "importer": "ExcelImporter",
            "supported_extensions": self.supported_extensions,
            "pandas_available": self._pandas_available,
            "openpyxl_available": self._openpyxl_available
        }
